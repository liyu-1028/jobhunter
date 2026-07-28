#!/usr/bin/env python3
"""导入教育部《全国普通高等学校名单》Excel 到本地高校名录.

用法:
    python scripts/import_moe_list.py <教育部名单.xlsx>

Excel 下载: 教育部官网 → 信息公开 → 高等学校名单
    https://www.moe.gov.cn/jyb_xxgk/s5743/s5744/  [需验证最新地址]

仅离线导入依赖 openpyxl (pip install openpyxl), 运行期不依赖.
产出:
    data/moe_universities.json  —— 运行期注册表读取的名录快照
    data/jobhunter.db universities 表 —— 同份数据的 SQLite 镜像
"""

import json
import os
import re
import sqlite3
import sys

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
JSON_OUT = os.path.join(PROJECT_ROOT, "data", "moe_universities.json")
DB_PATH = os.path.join(PROJECT_ROOT, "data", "jobhunter.db")


def parse_region(location: str):
    """拆分"所在地"为 (省份, 城市): '安徽省合肥市' → ('安徽', '合肥'), '北京市' → ('北京', '北京')"""
    loc = (location or "").strip()
    m = re.match(r"^(.+?(?:省|自治区|市))(.*)$", loc)
    if not m:
        return (loc or "unknown"), ""
    province = re.sub(r"(省|自治区|市)$", "", m.group(1))
    city = re.sub(r"(市|地区|自治州|盟)$", "", m.group(2)) or province
    return province, city


def _cell(row, idx):
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def main(excel_path: str):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("❌ 缺少依赖 openpyxl，请先运行: pip install openpyxl")

    if not os.path.exists(excel_path):
        sys.exit(f"❌ 文件不存在: {excel_path}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    # 定位表头行 (含"学校名称"的行)
    header_idx = None
    for i, row in enumerate(rows[:15]):
        if any(cell and "学校名称" in str(cell) for cell in row):
            header_idx = i
            break
    if header_idx is None:
        sys.exit("❌ 未找到表头行 (应包含「学校名称」)，请确认使用的是教育部名单 Excel")

    header = [str(c).strip() if c else "" for c in rows[header_idx]]

    def col(*keywords):
        for j, name in enumerate(header):
            if any(k in name for k in keywords):
                return j
        return None

    c_name = col("学校名称")
    c_auth = col("主管部门")
    c_loc = col("所在地")
    c_level = col("招生层次", "办学层次", "层次")
    c_owner = col("办学性质", "性质")

    records, seen = [], set()
    for row in rows[header_idx + 1:]:
        name = _cell(row, c_name)
        if not name or name in seen:
            continue
        seen.add(name)
        province, city = parse_region(_cell(row, c_loc))
        records.append({
            "name": name,
            "authority": _cell(row, c_auth),
            "province": province,
            "city": city,
            "level": _cell(row, c_level),
            "ownership": _cell(row, c_owner),
        })

    if not records:
        sys.exit("❌ 未解析到任何高校记录，请检查 Excel 格式")

    # 写出 JSON 快照 (运行期注册表数据源)
    os.makedirs(os.path.dirname(JSON_OUT), exist_ok=True)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    # SQLite 镜像
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DROP TABLE IF EXISTS universities")
    conn.execute("""
    CREATE TABLE universities (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        authority TEXT,
        province TEXT,
        city TEXT,
        level TEXT,
        ownership TEXT,
        updated_at TEXT
    )
    """)
    for r in records:
        conn.execute(
            """INSERT OR IGNORE INTO universities
               (name, authority, province, city, level, ownership, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, datetime('now', 'localtime'))""",
            (r["name"], r["authority"], r["province"], r["city"], r["level"], r["ownership"]),
        )
    conn.commit()
    conn.close()

    print(f"✅ 已导入 {len(records)} 所高校")
    print(f"   JSON 快照: {JSON_OUT}")
    print(f"   SQLite 镜像: {DB_PATH} (universities 表)")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(
            "用法: python scripts/import_moe_list.py <教育部高校名单.xlsx>\n"
            "下载: 教育部官网 → 信息公开 → 高等学校名单"
        )
    main(sys.argv[1])
